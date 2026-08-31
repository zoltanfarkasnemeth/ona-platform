# -*- coding: utf-8 -*-
"""
ONA párosító motor
==================

Beolvassa a db/ona.db adatbázis két tábláját (keresok, munkavallalok), és
minden olyan (kereső, munkavállaló) párost, ahol:
  1) azonos a szolgáltatási kategória (Senior / Recovery / Mama / Care), ÉS
  2) a munkavállaló havi díjigénye legfeljebb 20%-kal tér el a kereső havi
     keretétől (a két érték átlagához viszonyítva),
kiír a matches/talalatok.xlsx táblázatba.

Az eltérés% oszlop Excel-képlet, tehát a fájlban közvetlenül ellenőrizhető /
újraszámolható: =ABS(munkavallalo_dij - kereso_keret) / ((munkavallalo_dij + kereso_keret) / 2)

Futtatás:
    python3 tools/match_engine.py

Újrafuttatható: minden futáskor a jelenlegi adatbázis-tartalom alapján teljes
egészében újraírja a talalatok.xlsx-et.
"""
import os
import sqlite3
import sys
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
DB_PATH = os.path.join(ROOT, "db", "ona.db")
OUT_PATH = os.path.join(ROOT, "matches", "talalatok.xlsx")

MATCH_THRESHOLD_PCT = 20.0  # a megengedett eltérés a két oldal fizetési igénye között


def load_rows():
    if not os.path.exists(DB_PATH):
        sys.exit(f"Nem található az adatbázis: {DB_PATH}. Előbb futtasd: python3 db/seed_data.py")
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    keresok = conn.execute("SELECT * FROM keresok ORDER BY id").fetchall()
    munkavallalok = conn.execute("SELECT * FROM munkavallalok ORDER BY id").fetchall()
    conn.close()
    return keresok, munkavallalok


def find_matches(keresok, munkavallalok):
    """Python oldalon dönti el, mely párok kerülnek a táblázatba (kategória +
    20%-os fizetési sáv). Az Excelbe kerülő eltérés%-ot maga a munkafüzet is
    újraszámolja képlettel, ez csak a szűréshez kell."""
    matches = []
    for k in keresok:
        for m in munkavallalok:
            if k["kategoria"] != m["kategoria"]:
                continue
            keret = k["havi_keret_ft"]
            dij = m["havi_dij_igeny_ft"]
            atlag = (keret + dij) / 2
            if atlag == 0:
                continue
            elteres_pct = abs(dij - keret) / atlag * 100
            if elteres_pct <= MATCH_THRESHOLD_PCT:
                matches.append((k, m, elteres_pct))
    # legjobb egyezés (legkisebb eltérés) elöl
    matches.sort(key=lambda row: row[2])
    return matches


HEADERS = [
    "Kereső neve", "Kereső email", "Kereső telefon", "Kereső települése",
    "Kategória",
    "Kereső havi kerete (Ft)",
    "Munkavállaló neve", "Munkavállaló email", "Munkavállaló telefon", "Munkavállaló települése",
    "Munkavállaló havi díjigénye (Ft)",
    "Eltérés (%)",
    "Állapot",
    "Párosítás dátuma",
]

MONEY_FMT = '#,##0" Ft"'
PCT_FMT = "0.0%"

NAVY = "132447"
GOLD = "C68A3E"
LIGHT = "F7F8FB"


def build_workbook(matches):
    wb = Workbook()
    ws = wb.active
    ws.title = "Találatok"

    ws.append(HEADERS)
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor=NAVY)
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 34

    today_str = date.today().isoformat()
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, (k, m, _elteres_pct) in enumerate(matches, start=2):
        ws.cell(row=i, column=1, value=k["nev"])
        ws.cell(row=i, column=2, value=k["email"])
        ws.cell(row=i, column=3, value=k["telefon"])
        ws.cell(row=i, column=4, value=k["telepules"])
        ws.cell(row=i, column=5, value=k["kategoria"])
        ws.cell(row=i, column=6, value=k["havi_keret_ft"]).number_format = MONEY_FMT
        ws.cell(row=i, column=7, value=m["nev"])
        ws.cell(row=i, column=8, value=m["email"])
        ws.cell(row=i, column=9, value=m["telefon"])
        ws.cell(row=i, column=10, value=m["telepules"])
        ws.cell(row=i, column=11, value=m["havi_dij_igeny_ft"]).number_format = MONEY_FMT
        # Eltérés% élő Excel-képlettel: |K - F| / ((K + F) / 2), százalék formátumban (tört alakban tárolva)
        formula = f"=ABS(K{i}-F{i})/((K{i}+F{i})/2)"
        ws.cell(row=i, column=12, value=formula).number_format = PCT_FMT
        ws.cell(row=i, column=13, value="Új")
        ws.cell(row=i, column=14, value=today_str)
        for col_idx in range(1, len(HEADERS) + 1):
            ws.cell(row=i, column=col_idx).border = border
            ws.cell(row=i, column=col_idx).font = Font(name="Arial", size=10.5)
            if col_idx in (6, 11, 12):
                ws.cell(row=i, column=col_idx).alignment = Alignment(horizontal="right")

    # oszlopszélességek
    widths = [18, 26, 15, 22, 11, 16, 18, 26, 15, 22, 18, 12, 10, 14]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    ws.freeze_panes = "A2"

    last_row = max(len(matches) + 1, 2)
    if matches:
        tab = Table(displayName="Talalatok", ref=f"A1:N{last_row}")
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False
        )
        ws.add_table(tab)

    # magyarázó jegyzet a táblázat alá
    note_row = last_row + 3
    note_cell = ws.cell(
        row=note_row, column=1,
        value=(
            f"Párosítási szabály: azonos kategória, és az Eltérés (%) legfeljebb {MATCH_THRESHOLD_PCT:.0f}% "
            f"(|munkavállalói díjigény − kereső kerete| / átlaguk). Forrás: db/ona.db · Generálva: {today_str} "
            f"a tools/match_engine.py futtatásával."
        ),
    )
    note_cell.font = Font(name="Arial", size=9, italic=True, color="7A7A7A")
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=len(HEADERS))

    if not matches:
        empty_cell = ws.cell(row=2, column=1, value="Jelenleg nincs a feltételeknek megfelelő párosítás az adatbázisban.")
        empty_cell.font = Font(name="Arial", italic=True, color="7A7A7A")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    wb.save(OUT_PATH)


def main():
    keresok, munkavallalok = load_rows()
    matches = find_matches(keresok, munkavallalok)
    build_workbook(matches)
    print(f"{len(matches)} párosítás mentve ide: {OUT_PATH}")
    for k, m, pct in matches:
        print(f"  [{k['kategoria']}] {k['nev']} ({k['havi_keret_ft']:,} Ft) <-> {m['nev']} ({m['havi_dij_igeny_ft']:,} Ft) — eltérés {pct:.1f}%")


if __name__ == "__main__":
    main()
